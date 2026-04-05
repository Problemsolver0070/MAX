"""Document tools — PDF, spreadsheet, CSV, and JSON handling."""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from max.tools.registry import ToolDefinition

# ── Optional dependency imports ───────────────────────────────────────

try:
    from PyPDF2 import PdfReader

    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    from openpyxl import Workbook, load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from jsonpath_ng import parse as jsonpath_parse

    HAS_JSONPATH = True
except ImportError:
    HAS_JSONPATH = False

# ── Tool definitions ──────────────────────────────────────────────────

TOOL_DEFINITIONS = [
    ToolDefinition(
        tool_id="document.read_pdf",
        category="document",
        description="Extract text from a PDF file. Optionally specify page range.",
        permissions=["fs.read"],
        provider_id="native",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Absolute path to PDF file"},
                "pages": {
                    "type": "string",
                    "description": "Page range, e.g. '1-5' or '3' (1-based)",
                },
            },
            "required": ["path"],
        },
    ),
    ToolDefinition(
        tool_id="document.read_spreadsheet",
        category="document",
        description="Read an Excel or CSV file and return rows as JSON.",
        permissions=["fs.read"],
        provider_id="native",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Absolute path to Excel/CSV file"},
                "sheet": {
                    "type": "string",
                    "description": "Sheet name (Excel only, defaults to active sheet)",
                },
            },
            "required": ["path"],
        },
    ),
    ToolDefinition(
        tool_id="document.write_csv",
        category="document",
        description="Write rows of data to a CSV file.",
        permissions=["fs.write"],
        provider_id="native",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Absolute path for output CSV"},
                "rows": {
                    "type": "array",
                    "items": {"type": "object"},
                    "description": "List of row dicts to write",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column order (defaults to keys from first row)",
                },
            },
            "required": ["path", "rows"],
        },
    ),
    ToolDefinition(
        tool_id="document.write_spreadsheet",
        category="document",
        description="Write rows of data to an Excel (.xlsx) file.",
        permissions=["fs.write"],
        provider_id="native",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Absolute path for output Excel file"},
                "rows": {
                    "type": "array",
                    "items": {"type": "object"},
                    "description": "List of row dicts to write",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column order (defaults to keys from first row)",
                },
                "sheet": {
                    "type": "string",
                    "description": "Sheet name (defaults to 'Sheet')",
                },
            },
            "required": ["path", "rows"],
        },
    ),
    ToolDefinition(
        tool_id="document.parse_json",
        category="document",
        description="Parse a JSON file and optionally run a JSONPath query.",
        permissions=["fs.read"],
        provider_id="native",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Absolute path to JSON file"},
                "query": {
                    "type": "string",
                    "description": "JSONPath expression (requires jsonpath-ng)",
                },
            },
            "required": ["path"],
        },
    ),
]

# ── Page range parser ─────────────────────────────────────────────────


def _parse_page_range(pages_str: str, total_pages: int) -> list[int]:
    """Parse a page range string into a list of 0-based page indices.

    Accepts '3' (single page) or '2-5' (inclusive range). Pages are 1-based
    in the input and converted to 0-based indices internally.
    """
    pages_str = pages_str.strip()
    if "-" in pages_str:
        parts = pages_str.split("-", 1)
        start = max(int(parts[0]) - 1, 0)
        end = min(int(parts[1]), total_pages)
        return list(range(start, end))
    else:
        idx = int(pages_str) - 1
        if 0 <= idx < total_pages:
            return [idx]
        return []


# ── Handlers ──────────────────────────────────────────────────────────


async def handle_document_read_pdf(inputs: dict[str, Any]) -> dict[str, Any]:
    """Extract text from a PDF file."""
    if not HAS_PYPDF2:
        raise RuntimeError(
            "PyPDF2 is not installed. Install with: pip install 'max[documents]'"
        )

    path = Path(inputs["path"])
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    reader = PdfReader(str(path))
    total_pages = len(reader.pages)

    if "pages" in inputs and inputs["pages"]:
        page_indices = _parse_page_range(inputs["pages"], total_pages)
    else:
        page_indices = list(range(total_pages))

    text_parts: list[str] = []
    for idx in page_indices:
        page_text = reader.pages[idx].extract_text() or ""
        text_parts.append(page_text)

    return {
        "text": "\n".join(text_parts),
        "page_count": total_pages,
        "pages_read": len(page_indices),
    }


async def handle_document_read_spreadsheet(inputs: dict[str, Any]) -> dict[str, Any]:
    """Read an Excel or CSV file to JSON rows."""
    path = Path(inputs["path"])
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    suffix = path.suffix.lower()

    if suffix == ".csv":
        return _read_csv(path)
    elif suffix in {".xlsx", ".xls"}:
        return _read_excel(path, inputs.get("sheet"))
    else:
        raise ValueError(f"Unsupported file format: {suffix}. Use .csv, .xlsx, or .xls")


def _read_csv(path: Path) -> dict[str, Any]:
    """Read a CSV file and return rows as dicts."""
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        columns = reader.fieldnames or []
        rows = list(reader)
    return {"rows": rows, "columns": list(columns), "row_count": len(rows)}


def _read_excel(path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    """Read an Excel file and return rows as dicts."""
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is not installed. Install with: pip install 'max[documents]'"
        )

    wb = load_workbook(str(path), read_only=True, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        return {"rows": [], "columns": [], "row_count": 0}

    columns = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
    rows: list[dict[str, Any]] = []
    for row_values in rows_iter:
        row_dict = {}
        for col_name, value in zip(columns, row_values):
            row_dict[col_name] = value
        rows.append(row_dict)

    wb.close()
    return {"rows": rows, "columns": columns, "row_count": len(rows)}


async def handle_document_write_csv(inputs: dict[str, Any]) -> dict[str, Any]:
    """Write data rows to a CSV file."""
    path = Path(inputs["path"])
    path.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = inputs["rows"]
    if not rows:
        path.write_text("", encoding="utf-8")
        return {"rows_written": 0, "path": str(path)}

    columns = inputs.get("columns") or list(rows[0].keys())

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    return {"rows_written": len(rows), "path": str(path)}


async def handle_document_write_spreadsheet(inputs: dict[str, Any]) -> dict[str, Any]:
    """Write data rows to an Excel file."""
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is not installed. Install with: pip install 'max[documents]'"
        )

    path = Path(inputs["path"])
    path.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = inputs["rows"]
    sheet_name = inputs.get("sheet", "Sheet")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        wb.save(str(path))
        wb.close()
        return {"rows_written": 0, "path": str(path)}

    columns = inputs.get("columns") or list(rows[0].keys())

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))

    wb.save(str(path))
    wb.close()
    return {"rows_written": len(rows), "path": str(path)}


async def handle_document_parse_json(inputs: dict[str, Any]) -> dict[str, Any]:
    """Parse a JSON file and optionally apply a JSONPath query."""
    path = Path(inputs["path"])
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    text = path.read_text(encoding="utf-8")
    data = json.loads(text)

    query = inputs.get("query")
    if query:
        if not HAS_JSONPATH:
            raise RuntimeError(
                "jsonpath-ng is not installed. Install with: pip install 'max[documents]'"
            )
        expr = jsonpath_parse(query)
        matches = expr.find(data)
        return {"results": [m.value for m in matches]}

    return {"data": data}
